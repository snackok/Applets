
import os
from tkinter import *
from tkinter import ttk
from tkinter import filedialog
from tkinter import messagebox
from openpyxl import *

del_count = 0
fun_count = 0
first_path = ''
_xls_path = ''
all_file = []


class Application(Frame):
    '''一个经典的GUI程序的类写法'''

    def __init__(self, master=None):
        super().__init__(master)
        self.master = master
        self.pack()
        self.create_widget()

    def create_widget(self):
        global path
        path = StringVar()
        del_type = StringVar()
        del_type.set(1)

        # 创建组建
        Label(self, text="Excel路径:").grid(row=0, column=0)
        Entry(self, textvariable=path, width=25).grid(row=0, column=1, sticky=W)
        Button(self, text="路径选择", command=sel_path).grid(row=0, column=2, sticky=E)
        Button(self, text='打开表格', command=lambda: open_xls(path.get())).grid(row=1, columnspan=3)
        # 调试用
        path.set(r'W:/09TEMP/test.xlsx')


def sel_path():
        path_ = filedialog.askopenfilename()
        path.set(path_)


def open_xls(xls_path=os.getcwd()):
    if not os.path.exists(xls_path):
        messagebox.showerror('错误', '文件路径错误')
        return
    # print_info('打开文件：' + xls_path)
    global wb, titles, _xls_path, openwin
    openwin = []
    _xls_path = xls_path
    wb = load_workbook(xls_path)
    ws = wb.active
    # print_info(ws.title)

    mrow = ws.max_row
    mcol = ws.max_column

    xlarr = []
    for r in ws.rows:
        arr = []
        for cell in r:
            arr.append(cell.value)
        xlarr.append(arr)
    col_name = []
    for c in range(mcol):
        col_name.append(c)
    global form_list
    form_list = Tk()
    form_list.geometry('1000x400')
    global tree
    tree = ttk.Treeview(form_list, show='headings', columns=col_name, selectmode='browse')  # 单行选中模式

    for c in range(mcol):
        tree.column(c, width=90, anchor='center')
        tree.heading(c, text=xlarr[0][c])  # 显示标题
    titles = xlarr[0]
    del(xlarr[0])  # 删了第一个,不删显示内容第一行与标题重复
    for i in range(mrow-1):
        tree.insert('', i, values=xlarr[i], tags='oddrow')  # 显示内容
    tree.pack(side=TOP, fill=None)
    btn1 = Button(form_list, text='调试', command=save_book)
    btn1.pack(side=BOTTOM, expand=YES)

    # 设置单元格背景色
    tree.tag_configure('oddrow', background='orange')
    tree.bind('<Double-Button-1>', viewclick)
    # 关闭打开对话框
    # global root
    # oot.destroy()
    form_list.mainloop()

    # 把excel内容打印在控制台
    # for j in ws.rows:  # we.rows 获取每一行数据
    #     for n in j:
    #         print(n.value, end="\t")  # n.value 获取单元格的值
    #     print()
    # wb.close()


def viewclick(event):
    global nwin
    global tree
    global enty
    global sitem
    global colint
    for item in tree.selection():
        ttext = tree.item(item, 'values')
        sitem = item

    col = tree.identify_column(event.x)
    colint = int(str(col.replace('#', '')))

    nwin = Tk()  # 编辑窗口
    nwin.geometry("260x100")
    label1 = Label(nwin, text="修改：")
    label1.pack(side=LEFT, fill=None)
    enty = Text(nwin, width=300, height=300, wrap=WORD)
    enty = Entry(nwin)
    enty.pack(side=LEFT, fill=None)
    btn = Button(nwin, text='确认', command=getv)
    btn.pack(side=LEFT, padx=6, ipadx=6)
    enty.insert('end', ttext[colint - 1])  # 编辑框显示值

    openwin.append(nwin)

    if len(openwin) > 1:
        openwin.pop(0).destroy()
    nwin.protocol('WM_DELETE_WINDOW', initopenwin)  # 绑定时件,关闭窗清除变量值
    nwin.mainloop()


def initopenwin():
    global openwin
    openwin = []
    print_info('close')


def getv():
    global nwin
    global enty
    global tree
    global sitem
    global colint
    global openwin
    editxt = enty.get()
    tree.set(sitem, (colint - 1), editxt)
    openwin = []
    nwin.destroy()


def save_book():
    global wb
    global tree
    global titles
    global _xls_path

    ws = wb.create_sheet('change1')
    ws.append(titles)
    for itm in tree.get_children():
        ws.append(tree.item(itm)['values'])
    wb.save(_xls_path)
    messagebox.showinfo('提示', '保存成功')


def print_info(msg='xxx'):
    # 清理entry2
    messagebox.showinfo('提升', msg)


if __name__ == '__main__':
    # 初始化Tk()
    root = Tk()
    # form_list = root
    # 设置标题
    root.title('打开表格 by WW.WCGS')
    app = Application(master=root)
    root.mainloop()

