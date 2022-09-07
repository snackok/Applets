
import os
from tkinter import *
from tkinter import ttk
from tkinter import filedialog
import tkinter.messagebox
from openpyxl import *

del_count = 0
fun_count = 0
first_path = ''
all_file = []


def open_xls(xls_path=os.getcwd()):
    if not os.path.exists(xls_path):
        # print("文件夹路径错误")
        tkinter.messagebox.showerror('错误', '文件路径错误')
        return
    print_info('打开文件：' + xls_path)
    wb = load_workbook(xls_path)
    ws = wb.active
    print_info(ws.title)

    mrow = ws.max_row
    mcol = ws.max_column

    xlarr = []
    for r in ws.rows:
        arr = []
        for cell in r:
            arr.append(cell.value)
        xlarr.append(arr)

    form_list = Tk()
    form_list.geometry('800x400')

    col_name = []
    for c in range(mcol):
        col_name.append(c)

    tree = ttk.Treeview(form_list, show='headings', columns=col_name, selectmode='browse')  # 单行选中模式

    for c in range(mcol):
        tree.column(c, width=90, anchor='center')
        tree.heading(c, text=xlarr[0][c])  # 显示标题

    titles = xlarr[0]

    del(xlarr[0])  # 删了第一个,不删显示内容第一行与标题重复
    for i in range(mrow-1):
        tree.insert('', i, values=xlarr[i])  # 显示内容
    tree.pack(side=TOP, fill=None)
    btn1 = Button(form_list, text='保存', command=save_book)
    btn1.pack(side=BOTTOM, expand=YES)

    # 关闭打开对话框
    global root
    root.destroy()
    form_list.mainloop()

    # 把excel内容打印在控制台
    # for j in ws.rows:  # we.rows 获取每一行数据
    #     for n in j:
    #         print(n.value, end="\t")  # n.value 获取单元格的值
    #     print()
    # wb.close()


def save_book():
    tkinter.messagebox.showinfo('保存待写', msg)


def sel_path():
    path_ = filedialog.askopenfilename()
    path.set(path_)


def print_info(msg='xxx'):
    # 清理entry2
    tkinter.messagebox.showinfo('提升', msg)


if __name__ == '__main__':
    # 初始化Tk()
    root = Tk()
    # 设置标题
    root.title('打开表格 by WW.WCGS')
    path = StringVar()
    del_type = StringVar()
    del_type.set(1)

    Label(root, text="Excel路径:").grid(row=0, column=0)
    Entry(root, textvariable=path, width=25).grid(row=0, column=1, sticky=W)
    Button(root, text="路径选择", command=sel_path).grid(row=0, column=2, sticky=E)
    Button(root, text='打开表格', command=lambda: open_xls(path.get())).grid(row=1, columnspan=3)
    # 调试用
    path.set(r'W:/09TEMP/test.xlsx')
    root.mainloop()

