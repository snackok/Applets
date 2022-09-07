# -*- encoding:utf-8 -*-
"""
@作者：WW.WCGS,YYL.WCGS
@文件名：evluate.py
@时间：2022/1/12  13:23
@文档说明:
"""

import os
# import tkinter
import random
import time
from tkinter import *  # 引入模块中的所有‘公开’成员
from tkinter import ttk  # 引入模块中的ttk类(函数或者变量)
from tkinter import filedialog
from tkinter import messagebox
from openpyxl import *

# 全局变量
_xls_path = ''  # 测试数据路径
_titles = []  # 数据列头
_list_evlu_GB = []  # 评价标准数据
_openwin = []  # 打开的窗口
_xls_wb = None  # 分别为国标和评价xls的workbook对象
_tree = None  # TreeView控件
_form_list = None  # 主窗口

# 静态常量
c_gb_filename = '评价标准.xlsx'  # 国标excel 文件
c_debug_xls_name = r'W:/09TEMP/01/test.xlsx'  # 调试使用


# 主应用类
class Application(Frame):
    # 一个经典的GUI程序的类写法

    def __init__(self, master=None):
        super().__init__(master)
        self.master = master
        self.pack()
        self.create_widget()

    def create_widget(self):
        global _path
        _path = StringVar()

        # 创建组建
        Label(self, text="Excel路径:").grid(row=0, column=0)
        Entry(self, textvariable=_path, width=25).grid(row=0, column=1, sticky=W)
        Button(self, text="路径选择", command=sel_path).grid(row=0, column=2, sticky=E)
        Button(self, text='打开表格', command=lambda: open_xls(_path.get())).grid(row=1, columnspan=3)
        # 调试用
        global c_debug_xls_name
        _path.set(c_debug_xls_name)


def sel_path():
    m_xls_path = filedialog.askopenfilename()
    _path.set(m_xls_path)


# 打开xls文件，并显示在列表上
def open_xls(xls_path=os.getcwd()):
    if not os.path.exists(xls_path):
        messagebox.showerror('错误', '文件路径错误')
        return
    # print_info('打开文件：' + xls_path)
    global _xls_path, _xls_wb, _titles, _openwin
    _openwin = []
    _xls_path = xls_path
    m_xlarr = xls_2_list(xls_path, True)  # 获取评价数据，同时将_xls_wb赋值
    m_ws = _xls_wb.active

    m_row = m_ws.max_row
    m_col = m_ws.max_column

    m_col_name = []

    for c in range(m_col):
        m_col_name.append(c)
    global _form_list  # 表格窗体
    _form_list = Tk()
    _form_list.geometry('1000x400')
    global _tree

    # 变换treeview 颜色尝试，OK 1.14

    # 针对python3.9 不能更改TreeView 颜色的补丁。
    def fixed_map(option):
        return [elm for elm in m_style.map("Treeview", query_opt=option) if elm[:2] != ("!disabled", "!selected")]
    m_style = ttk.Style(_form_list)
    m_style.map('Treeview', foreground=fixed_map('foreground'), background=fixed_map('background'))
    # 补丁结束
    _tree = ttk.Treeview(_form_list, show='headings', columns=m_col_name, selectmode='browse')  # 单行选中模式
    for c in range(m_col):
        _tree.column(c, width=90, anchor='center')
        _tree.heading(c, text=m_xlarr[0][c])  # 显示标题
    _titles = m_xlarr[0]
    print("_titles:" + str(_titles))
    # del(m_xlarr[0])  # 删了第一个,不删显示内容第一行与标题重复
    for i in range(m_row - 1):
        _tree.insert('', i, values=m_xlarr[i + 1], tags='0')  # 显示内容
    _tree.pack(side=TOP, fill=None)

    btn1 = Button(_form_list, text='   评价   ', command=lambda: evl_2_list(m_xlarr))
    btn1.pack(side=BOTTOM, expand=YES)
    btn2 = Button(_form_list, text='   保存   ', command=save_book)
    btn2.pack(side=BOTTOM, expand=YES)
    btn3 = Button(_form_list, text='   测试   ', command=debug_btn)
    btn3.pack(side=BOTTOM)

    _tree.bind('<Double-Button-1>', viewclick)
    # 关闭打开对话框
    global _root
    _root.destroy()

    _form_list.mainloop()

    # 把excel内容打印在控制台
    # for j in ws.rows:  # we.rows 获取每一行数据
    #     for n in j:
    #         print(n.value, end="\t")  # n.value 获取单元格的值
    #     print()
    # wb.close()


def viewclick(event):
    global nwin
    global _tree
    global enty
    global sitem
    global colint
    for item in _tree.selection():
        ttext = _tree.item(item, 'values')
        sitem = item

    col = _tree.identify_column(event.x)
    colint = int(str(col.replace('#', '')))

    nwin = Tk()  # 编辑窗口
    nwin.geometry("260x100")
    label1 = Label(nwin, text="修改：")
    label1.pack(side=LEFT, fill=None)
    enty = Text(nwin, width=300, height=300, wrap=WORD)
    enty = Entry(nwin)
    enty.pack(side=LEFT, fill=None)
    btn = Button(nwin, text='确认', command=getv)
    btn.pack(side=LEFT, padx=6, ipadx=6)
    enty.insert('end', ttext[colint - 1])  # 编辑框显示值

    _openwin.append(nwin)

    if len(_openwin) > 1:
        _openwin.pop(0).destroy()
    nwin.protocol('WM_DELETE_WINDOW', close_win)  # 绑定时件,关闭窗清除变量值
    nwin.mainloop()


def debug_btn():
    print(f'\n\n')
    print(f'-------------------------------------------------调试信息-----------------------------------------------')
    print(f'----------------------测试TreeView中Item的颜色设置--------------------------')
    _tree.tag_configure('0', background="#C0C0C0", foreground="#ff0000")
    _tree.tag_configure('6', background="#C0C0C0", foreground="red")
    _tree.tag_configure('0', background="#C0C0C0", font="Arial 10")

    print(f'I005元素的value为：{_tree.item("I005", option="values")}')
    print(f'I005元素的tags为：{_tree.item("I005", option="tags")}')
    print(f'tags为0的背景色为：{_tree.tag_configure("0", option="foreground")}')
    # print(f'----------------------测试TreeView中对任意行列赋值--------------------------')
    # print(_tree.item('I005',value=['abc','def','ghi','jkl']))
    # print(_tree.set('I'+str(1).rjust(3,'0'),0,'xxx'))
    # print(f'------------------调试添加TreeView列--------------------------')
    # _tree.column("12", minwidth=0, width=100)
    # _tree.heading("12", text="添加")
    # _tree.update()
    # print(f"column：{ _tree['column']}heading:{_tree.heading('12',text = '测试')}")
    # _tree.column(c, width=90, anchor='center')
    # _tree.heading(c, text=m_xlarr[0][c])  # 显示标题
    print(f'-----------------------------------------------调试信息结束----------------------------------------------')
    print(f'\n\n')


# 关闭主窗口触发事件
def close_win():
    global _openwin, nwin
    _openwin = []
    nwin.destroy()


# 修改TreeView中的值
def getv():
    global nwin
    global enty
    global _tree
    global sitem
    global colint
    global _openwin
    editxt = enty.get()
    _tree.set(sitem, (colint - 1), editxt)
    _openwin = []
    nwin.destroy()


# 保存表格
def save_book():
    global _xls_wb
    global _tree
    global _titles
    # global _xls_path

    ws = _xls_wb.create_sheet(time.strftime("%Y-%m-%d %Hh", time.localtime()))
    ws.append(_titles)
    for itm in _tree.get_children():
        ws.append(_tree.item(itm)['values'])
    _xls_wb.save(_xls_path)
    messagebox.showinfo('提示', '保存成功')


# 将评价结果显示到TreeView控件上
def evl_2_list(p_list):
    # print(f'参数为:{ p_list }')
    m_list = evluate(p_list)
    # 将m_list 的数据更新TreeView
    print(f'\n\n评价结果为:{len(m_list)}行{len(m_list[0])}列。\n内容为：{m_list}')
    for c in range(len(m_list[0])):  # 遍历行
        if c > 0:
            for r in range(len(m_list)):  # 遍历列
                # 依次将m_list赋值给treeList
                _tree.set('I' + str(c).rjust(3, '0'), r, m_list[r][c])
                t_cur_tag = _tree.item('I' + str(c).rjust(3, '0'), option='tags')
                print(f'---{type(m_list[r][c])}--{m_list[r][c]}------')
                # 如果遍历元素为数字，并且比这一行的之前设置的tags 大，则tags赋值为此数字
                if str(m_list[r][c]).isdigit() and int(str(t_cur_tag[0])) < int(m_list[r][c]):
                    _tree.item('I' + str(c).rjust(3, '0'), tags=str(m_list[r][c]))

    _tree.tag_configure('6', background="#C0C0C0", foreground="red")
    _tree.tag_configure('5', background="#C0C0C0", foreground="orange")
    _tree.tag_configure('4', background="#C0C0C0", foreground="blue")
    _tree.tag_configure('3', background="#C0C0C0", foreground="yellow")


# 评价所有列函数 , arr 为传入的表格数据，返回list
def evluate(p_list):
    # 取出评价标准  \\\\\\
    # print(_xls_path)
    global _list_evlu_GB
    _list_evlu_GB = xls_2_list(get_GB())  # 读取评价内容
    print("评价内容：" + str(_list_evlu_GB))
    re_list_evlu = []
    # 循环每一列进行评价
    for c in range(len(p_list[0])):
        m_col = [i[c] for i in p_list]  # 取第c列
        if c > 1:  # 除去送样号和分析号列
            # print(f"p_list数据第{c+1}列:{m_col},/n 评价结果：{ evlu_element(m_col)}")
            re_list_evlu.append(evlu_element(m_col))
        else:
            re_list_evlu.append(m_col)

    # print(f"p_list数据行:{len(p_list)},列{len(p_list[0])}")
    # m_first_col = [i[0] for i in p_list]  # 取第一列
    print("evluate结果为：\n" + str(re_list_evlu))
    return re_list_evlu


# 返回程序所在目录下的评价标准文件全名（带路径）
def get_GB():
    m_path = os.path.join(os.getcwd(), c_gb_filename)
    # print(f'get_GB函数：返回值为：{m_path}')
    # return r'W:/09TEMP/01/评价标准.xlsx'
    return m_path


# 将指定文件路径的xls数据读取后，返回数据list
# 如果p_is_xlswb 参数为true 表示打开的表格是评价数据，需要给_xls_wb赋值
def xls_2_list(p_xls_path, p_is_xlswb=False):
    if p_is_xlswb:
        global _xls_wb
        _xls_wb = load_workbook(p_xls_path)
        t_wb = _xls_wb
    else:
        t_wb = load_workbook(p_xls_path)
    t_ws = t_wb.active
    re_arr = []
    for r in t_ws.rows:
        t_arr = []
        for cell in r:
            t_arr.append(cell.value)
        re_arr.append(t_arr)  # 将表格存储到xlarr二维数组
    t_wb.close()
    # print(f'读入{p_xls_path}文件，输出列表：{re_arr}')
    return re_arr


# 评价某个元素,p_col 为传入的某元素列，第一个为元素名，第二个为单位，后面为数据。
# 返回list，将数据替换为评价级别。
def evlu_element(p_col):
    # print(f"评价列{p_col}")
    # 取出列头，即元素
    m_head = p_col[0]  # 列头，元素名
    m_unit = p_col[1]  # 第二列，单位
    # find_gb_list()  在_list_evlu_GB 列表中寻找对应元素列
    for r in range(len(p_col)):
        if r > 1:
            # 进行判别，获取类别
            p_col[r] = random.randint(1, 6)

    return p_col


def print_info(msg='xxx'):
    messagebox.showinfo('提示', msg)


# 主函数
if __name__ == '__main__':
    # 初始化Tk()
    _root = Tk()
    # form_list = root
    # 设置标题
    _root.title('打开表格 by WCGS')
    app = Application(master=_root)
    _root.mainloop()
